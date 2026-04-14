# -*- coding: utf-8 -*-
"""
Genera Excel con el calendario completo del Mundial 2026
Fuente: TYC Sports (fixture oficial FIFA)
Zona horaria destino: Chile (CLT = UTC-4 en junio/julio = invierno)
Conversión: hora Argentina (ART UTC-3) → Chile (CLT UTC-4) = ART - 1h
"""

import pandas as pd
from datetime import datetime, timezone, timedelta

# ──────────────────────────────────────────────────────────────────────────────
# DATOS COMPLETOS — FASE DE GRUPOS (fuente: TYC Sports, abril 2026)
# ──────────────────────────────────────────────────────────────────────────────
# Formato: (local, visita, fase, fecha_chile "YYYY-MM-DDTHH:MM", sede)
# fecha_chile: hora ya convertida a Santiago (UTC-4)

PARTIDOS_GRUPOS = [
    # ── GRUPO A ───────────────────────────────────────────────────────────────
    ("México",           "Sudáfrica",         "Grupo A", "2026-06-11T15:00", "Estadio Ciudad de México"),
    ("Corea del Sur",    "República Checa",   "Grupo A", "2026-06-11T22:00", "Estadio Guadalajara"),
    ("República Checa",  "Sudáfrica",         "Grupo A", "2026-06-18T12:00", "Atlanta Stadium"),
    ("México",           "Corea del Sur",     "Grupo A", "2026-06-18T21:00", "Estadio Guadalajara"),
    ("República Checa",  "México",            "Grupo A", "2026-06-24T21:00", "Estadio Ciudad de México"),
    ("Sudáfrica",        "Corea del Sur",     "Grupo A", "2026-06-24T21:00", "Estadio Monterrey"),

    # ── GRUPO B ───────────────────────────────────────────────────────────────
    ("Canadá",           "Bosnia",            "Grupo B", "2026-06-12T15:00", "Toronto Stadium"),
    ("Qatar",            "Suiza",             "Grupo B", "2026-06-13T15:00", "San Francisco Bay Area Stadium"),
    ("Suiza",            "Bosnia",            "Grupo B", "2026-06-18T15:00", "Los Angeles Stadium"),
    ("Canadá",           "Qatar",             "Grupo B", "2026-06-18T18:00", "BC Place Vancouver"),
    ("Suiza",            "Canadá",            "Grupo B", "2026-06-24T15:00", "BC Place Vancouver"),
    ("Bosnia",           "Qatar",             "Grupo B", "2026-06-24T15:00", "Seattle Stadium"),

    # ── GRUPO C ───────────────────────────────────────────────────────────────
    ("Brasil",           "Marruecos",         "Grupo C", "2026-06-13T18:00", "New York / New Jersey Stadium"),
    ("Haití",            "Escocia",           "Grupo C", "2026-06-13T21:00", "Boston Stadium"),
    ("Escocia",          "Marruecos",         "Grupo C", "2026-06-19T18:00", "Boston Stadium"),
    ("Brasil",           "Haití",             "Grupo C", "2026-06-19T21:00", "Philadelphia Stadium"),
    ("Escocia",          "Brasil",            "Grupo C", "2026-06-24T18:00", "Miami Stadium"),
    ("Marruecos",        "Haití",             "Grupo C", "2026-06-24T18:00", "Atlanta Stadium"),

    # ── GRUPO D ───────────────────────────────────────────────────────────────
    ("Estados Unidos",   "Paraguay",          "Grupo D", "2026-06-12T21:00", "Los Angeles Stadium"),
    ("Australia",        "Turquía",           "Grupo D", "2026-06-13T00:00", "BC Place Vancouver"),
    ("Turquía",          "Paraguay",          "Grupo D", "2026-06-19T00:00", "San Francisco Bay Area Stadium"),
    ("Estados Unidos",   "Australia",         "Grupo D", "2026-06-19T15:00", "Seattle Stadium"),
    ("Turquía",          "Estados Unidos",    "Grupo D", "2026-06-25T22:00", "Los Angeles Stadium"),
    ("Paraguay",         "Australia",         "Grupo D", "2026-06-25T22:00", "San Francisco Bay Area Stadium"),

    # ── GRUPO E ───────────────────────────────────────────────────────────────
    ("Alemania",         "Curazao",           "Grupo E", "2026-06-14T13:00", "Houston Stadium"),
    ("Costa de Marfil",  "Ecuador",           "Grupo E", "2026-06-14T19:00", "Philadelphia Stadium"),
    ("Alemania",         "Costa de Marfil",   "Grupo E", "2026-06-20T16:00", "Toronto Stadium"),
    ("Curazao",          "Ecuador",           "Grupo E", "2026-06-20T20:00", "Kansas City Stadium"),
    ("Ecuador",          "Alemania",          "Grupo E", "2026-06-25T16:00", "New York / New Jersey Stadium"),
    ("Curazao",          "Costa de Marfil",   "Grupo E", "2026-06-25T16:00", "Philadelphia Stadium"),

    # ── GRUPO F ───────────────────────────────────────────────────────────────
    ("Países Bajos",     "Japón",             "Grupo F", "2026-06-14T16:00", "Dallas Stadium"),
    ("Suecia",           "Túnez",             "Grupo F", "2026-06-14T22:00", "Estadio Monterrey"),
    ("Japón",            "Túnez",             "Grupo F", "2026-06-20T00:00", "Estadio Monterrey"),
    ("Países Bajos",     "Suecia",            "Grupo F", "2026-06-20T13:00", "Houston Stadium"),
    ("Túnez",            "Países Bajos",      "Grupo F", "2026-06-25T19:00", "Dallas Stadium"),
    ("Japón",            "Suecia",            "Grupo F", "2026-06-25T19:00", "Kansas City Stadium"),

    # ── GRUPO G ───────────────────────────────────────────────────────────────
    ("Bélgica",          "Egipto",            "Grupo G", "2026-06-15T15:00", "Seattle Stadium"),
    ("Irán",             "Nueva Zelanda",     "Grupo G", "2026-06-15T21:00", "Los Angeles Stadium"),
    ("Bélgica",          "Irán",              "Grupo G", "2026-06-21T15:00", "Los Angeles Stadium"),
    ("Egipto",           "Nueva Zelanda",     "Grupo G", "2026-06-21T21:00", "BC Place Vancouver"),
    ("Nueva Zelanda",    "Bélgica",           "Grupo G", "2026-06-26T23:00", "BC Place Vancouver"),
    ("Egipto",           "Irán",              "Grupo G", "2026-06-26T23:00", "Seattle Stadium"),

    # ── GRUPO H ───────────────────────────────────────────────────────────────
    ("España",           "Cabo Verde",        "Grupo H", "2026-06-15T12:00", "Atlanta Stadium"),
    ("Arabia Saudita",   "Uruguay",           "Grupo H", "2026-06-15T18:00", "Miami Stadium"),
    ("España",           "Arabia Saudita",    "Grupo H", "2026-06-21T12:00", "Atlanta Stadium"),
    ("Cabo Verde",       "Uruguay",           "Grupo H", "2026-06-21T18:00", "Miami Stadium"),
    ("Uruguay",          "España",            "Grupo H", "2026-06-26T20:00", "Estadio Guadalajara"),
    ("Cabo Verde",       "Arabia Saudita",    "Grupo H", "2026-06-26T20:00", "Houston Stadium"),

    # ── GRUPO I ───────────────────────────────────────────────────────────────
    ("Francia",          "Senegal",           "Grupo I", "2026-06-16T15:00", "New York / New Jersey Stadium"),
    ("Irak",             "Noruega",           "Grupo I", "2026-06-16T18:00", "Boston Stadium"),
    ("Francia",          "Irak",              "Grupo I", "2026-06-22T17:00", "Philadelphia Stadium"),
    ("Noruega",          "Senegal",           "Grupo I", "2026-06-22T20:00", "New York / New Jersey Stadium"),
    ("Noruega",          "Francia",           "Grupo I", "2026-06-26T15:00", "Boston Stadium"),
    ("Senegal",          "Irak",              "Grupo I", "2026-06-26T15:00", "Toronto Stadium"),

    # ── GRUPO J ───────────────────────────────────────────────────────────────
    ("Austria",          "Jordania",          "Grupo J", "2026-06-16T00:00", "San Francisco Bay Area Stadium"),
    ("Argentina",        "Argelia",           "Grupo J", "2026-06-16T21:00", "Kansas City Stadium"),
    ("Argentina",        "Austria",           "Grupo J", "2026-06-22T13:00", "Dallas Stadium"),
    ("Jordania",         "Argelia",           "Grupo J", "2026-06-22T23:00", "San Francisco Bay Area Stadium"),
    ("Jordania",         "Argentina",         "Grupo J", "2026-06-27T22:00", "Dallas Stadium"),
    ("Argelia",          "Austria",           "Grupo J", "2026-06-27T22:00", "Kansas City Stadium"),

    # ── GRUPO K ───────────────────────────────────────────────────────────────
    ("Portugal",         "RD Congo",          "Grupo K", "2026-06-17T13:00", "Houston Stadium"),
    ("Uzbekistán",       "Colombia",          "Grupo K", "2026-06-17T22:00", "Estadio Ciudad de México"),
    ("Portugal",         "Uzbekistán",        "Grupo K", "2026-06-23T13:00", "Houston Stadium"),
    ("RD Congo",         "Colombia",          "Grupo K", "2026-06-23T22:00", "Estadio Guadalajara"),
    ("Colombia",         "Portugal",          "Grupo K", "2026-06-27T19:30", "Miami Stadium"),
    ("RD Congo",         "Uzbekistán",        "Grupo K", "2026-06-27T19:30", "Atlanta Stadium"),

    # ── GRUPO L ───────────────────────────────────────────────────────────────
    ("Inglaterra",       "Croacia",           "Grupo L", "2026-06-17T16:00", "Dallas Stadium"),
    ("Ghana",            "Panamá",            "Grupo L", "2026-06-17T19:00", "Toronto Stadium"),
    ("Inglaterra",       "Ghana",             "Grupo L", "2026-06-23T16:00", "Boston Stadium"),
    ("Croacia",          "Panamá",            "Grupo L", "2026-06-23T19:00", "Toronto Stadium"),
    ("Panamá",           "Inglaterra",        "Grupo L", "2026-06-27T17:00", "New York / New Jersey Stadium"),
    ("Croacia",          "Ghana",             "Grupo L", "2026-06-27T17:00", "Philadelphia Stadium"),
]

# ──────────────────────────────────────────────────────────────────────────────
# FASES ELIMINATORIAS (equipos por definir)
# ──────────────────────────────────────────────────────────────────────────────
TBD = "Por definir"

PARTIDOS_ELIMINATORIAS = [
    # ── Ronda de 32 (16 partidos) — 1-4 julio 2026 ───────────────────────────
    (TBD, TBD, "Ronda de 32", "2026-07-01T13:00", "Sede por confirmar"),
    (TBD, TBD, "Ronda de 32", "2026-07-01T17:00", "Sede por confirmar"),
    (TBD, TBD, "Ronda de 32", "2026-07-01T21:00", "Sede por confirmar"),
    (TBD, TBD, "Ronda de 32", "2026-07-02T13:00", "Sede por confirmar"),
    (TBD, TBD, "Ronda de 32", "2026-07-02T17:00", "Sede por confirmar"),
    (TBD, TBD, "Ronda de 32", "2026-07-02T21:00", "Sede por confirmar"),
    (TBD, TBD, "Ronda de 32", "2026-07-03T13:00", "Sede por confirmar"),
    (TBD, TBD, "Ronda de 32", "2026-07-03T17:00", "Sede por confirmar"),
    (TBD, TBD, "Ronda de 32", "2026-07-03T21:00", "Sede por confirmar"),
    (TBD, TBD, "Ronda de 32", "2026-07-04T13:00", "Sede por confirmar"),
    (TBD, TBD, "Ronda de 32", "2026-07-04T17:00", "Sede por confirmar"),
    (TBD, TBD, "Ronda de 32", "2026-07-04T21:00", "Sede por confirmar"),
    (TBD, TBD, "Ronda de 32", "2026-07-05T13:00", "Sede por confirmar"),
    (TBD, TBD, "Ronda de 32", "2026-07-05T17:00", "Sede por confirmar"),
    (TBD, TBD, "Ronda de 32", "2026-07-05T21:00", "Sede por confirmar"),
    (TBD, TBD, "Ronda de 32", "2026-07-06T17:00", "Sede por confirmar"),

    # ── Octavos de Final (8 partidos) — 9-12 julio ───────────────────────────
    (TBD, TBD, "Octavos", "2026-07-09T17:00", "Sede por confirmar"),
    (TBD, TBD, "Octavos", "2026-07-09T21:00", "Sede por confirmar"),
    (TBD, TBD, "Octavos", "2026-07-10T17:00", "Sede por confirmar"),
    (TBD, TBD, "Octavos", "2026-07-10T21:00", "Sede por confirmar"),
    (TBD, TBD, "Octavos", "2026-07-11T17:00", "Sede por confirmar"),
    (TBD, TBD, "Octavos", "2026-07-11T21:00", "Sede por confirmar"),
    (TBD, TBD, "Octavos", "2026-07-12T17:00", "Sede por confirmar"),
    (TBD, TBD, "Octavos", "2026-07-12T21:00", "Sede por confirmar"),

    # ── Cuartos de Final (4 partidos) — 16-17 julio ──────────────────────────
    (TBD, TBD, "Cuartos de Final", "2026-07-16T17:00", "Sede por confirmar"),
    (TBD, TBD, "Cuartos de Final", "2026-07-16T21:00", "Sede por confirmar"),
    (TBD, TBD, "Cuartos de Final", "2026-07-17T17:00", "Sede por confirmar"),
    (TBD, TBD, "Cuartos de Final", "2026-07-17T21:00", "Sede por confirmar"),

    # ── Semifinales (2 partidos) — 21-22 julio ───────────────────────────────
    (TBD, TBD, "Semifinal", "2026-07-21T21:00", "Sede por confirmar"),
    (TBD, TBD, "Semifinal", "2026-07-22T21:00", "Sede por confirmar"),

    # ── Tercer Puesto — 25 julio ─────────────────────────────────────────────
    (TBD, TBD, "Tercer Puesto", "2026-07-25T20:00", "Sede por confirmar"),

    # ── FINAL — 19 julio ─────────────────────────────────────────────────────
    (TBD, TBD, "Final", "2026-07-19T20:00", "New York / New Jersey Stadium"),
]


def build_dataframe() -> pd.DataFrame:
    rows = []
    for local, visita, fase, dt_chile, sede in PARTIDOS_GRUPOS + PARTIDOS_ELIMINATORIAS:
        dt = datetime.strptime(dt_chile, "%Y-%m-%dT%H:%M")
        dt_chile_tz = dt.replace(tzinfo=timezone(timedelta(hours=-4)))
        dt_utc = dt_chile_tz.astimezone(timezone.utc).strftime("%Y-%m-%dT%H:%M")
        rows.append({
            "local":            local,
            "visita":           visita,
            "torneo":           "Mundial 2026",
            "fase":             fase,
            "sede":             sede,
            "fecha_hora_chile": dt_chile,
            "fecha_hora_utc":   dt_utc,
        })
    return pd.DataFrame(rows)


def main():
    print("Generando Excel del Mundial 2026 (datos TYC Sports)...")

    df = build_dataframe()
    output = "mundial_2026_v2.xlsx"

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Partidos")
        ws = writer.sheets["Partidos"]
        for col, width in zip("ABCDEFG", [22, 22, 14, 18, 38, 22, 22]):
            ws.column_dimensions[col].width = width

        from openpyxl.styles import Font, PatternFill, Alignment
        header_fill = PatternFill("solid", fgColor="1E3A5F")
        for cell in ws[1]:
            cell.font      = Font(bold=True, color="FFFFFF")
            cell.fill       = header_fill
            cell.alignment  = Alignment(horizontal="center")

        tbd_fill = PatternFill("solid", fgColor="F0F0F0")
        for row in ws.iter_rows(min_row=2):
            if row[0].value == "Por definir":
                for cell in row:
                    cell.fill = tbd_fill

    grupos = df[df["fase"].str.startswith("Grupo")].shape[0]
    elim   = df[~df["fase"].str.startswith("Grupo")].shape[0]

    print(f"\nExcel guardado: {output}")
    print(f"  Partidos grupos:       {grupos}")
    print(f"  Partidos eliminatorio: {elim}")
    print(f"  TOTAL:                 {len(df)}")


if __name__ == "__main__":
    main()
